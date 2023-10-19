# views.py
from django.shortcuts import render
import json
from openpyxl import load_workbook
from django.http import JsonResponse
from rest_framework.decorators import api_view
from openpyxl.utils.dataframe import dataframe_to_rows
from rest_framework import status
import pandas as pd

@api_view(['POST'])
def excel_view(request, sheet_name):
    if request.method == 'POST':
        try:
            data = json.loads(request.body.decode("utf-8"))
            json_objects = data.get('json_objects')

            # Provide the full path to your Excel file
            file_path = r'C:\Users\99ksh\OneDrive\Documents\Desktop\PYEX_01\main.xlsx'

            if not json_objects:
                return JsonResponse({"error": "JSON objects are required."}, status=status.HTTP_400_BAD_REQUEST)

            # Load the Excel workbook using openpyxl
            workbook = load_workbook(filename=file_path)

            # Select the existing sheet or create a new one if it doesn't exist
            if sheet_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(title=sheet_name)
                sheet.append(['sr_no', 'name', 'amount', 'pending'])
            else:
                sheet = workbook[sheet_name]

            # Get the last sr_no in the sheet
            last_sr_no = max([row[0] for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True)], default=0)

            # Initialize sr_no to the next number
            sr_no = last_sr_no + 1

            # Iterate through JSON objects and update the Excel file
            for obj in json_objects:
                row = [sr_no, obj['name'], obj['amount'], obj['pending']]
                sheet.append(row)
                sr_no += 1

            # Save the updated Excel file
            workbook.save(file_path)

            return JsonResponse({'message': 'Data created successfully'}, status=status.HTTP_200_OK)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def read_excel(request, sheet_name):
    try:
        # Provide the full path to your Excel file
        file_path = r'C:\Users\99ksh\OneDrive\Documents\Desktop\PYEX_01\main.xlsx'

        # Load the Excel workbook using openpyxl
        workbook = load_workbook(filename=file_path)

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            excel_data = []

            # Extract data from the Excel sheet
            for row in sheet.iter_rows(min_row=2, values_only=True):
                sr_no, name, amount, pending = row
                excel_data.append({
                    "sr_no": sr_no,
                    "name": name,
                    "amount": amount,
                    "pending": pending
                })

            return JsonResponse({'data': excel_data}, status=status.HTTP_200_OK)
        else:
            return JsonResponse({'error': 'Sheet not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
@api_view(['PUT'])
def update_excel(request, sheet_name):
    if request.method == 'PUT':
        try:
            data = json.loads(request.body.decode("utf-8"))
            json_objects = data.get('json_objects')

            # Provide the full path to your Excel file
            file_path = r'C:\Users\99ksh\OneDrive\Documents\Desktop\PYEX_01\main.xlsx'

            if not json_objects:
                return JsonResponse({"error": "JSON objects are required."}, status=status.HTTP_400_BAD_REQUEST)

            # Load the Excel workbook using openpyxl
            workbook = load_workbook(filename=file_path)

            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Iterate through JSON objects and update the Excel file
                for obj in json_objects:
                    sr_no = obj.get('sr_no')
                    name = obj.get('name')
                    amount = obj.get('amount')
                    pending = obj.get('pending')

                    if sr_no is None:
                        return JsonResponse({'error': 'sr_no is mandatory for updates'}, status=status.HTTP_400_BAD_REQUEST)

                    found = False

                    # Find and update the row with the matching sr_no
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                        if row[0].value == sr_no:
                            row[1].value = name
                            row[2].value = amount
                            row[3].value = pending
                            found = True
                            break

                    if not found:
                        return JsonResponse({'error': f"Row with sr_no {sr_no} not found"}, status=status.HTTP_404_NOT_FOUND)

                # Save the updated Excel file
                workbook.save(file_path)

                return JsonResponse({'message': 'Data updated successfully'}, status=status.HTTP_200_OK)
            else:
                return JsonResponse({'error': 'Sheet not found'}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)





@api_view(['DELETE'])
def delete_excel(request, sheet_name):
    if request.method == 'DELETE':
        try:
            # Provide the full path to your Excel file
            file_path = r'C:\Users\99ksh\OneDrive\Documents\Desktop\PYEX_01\main.xlsx'
            
            # Load the Excel workbook using openpyxl
            workbook = load_workbook(filename=file_path)

            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Find the last row with data
                last_row = sheet.max_row

                if last_row > 1:
                    # Clear the data in the sheet, leaving only the header row
                    for row in sheet.iter_rows(min_row=2, max_row=last_row):
                        for cell in row:
                            cell.value = None

                # Save the updated Excel file
                workbook.save(file_path)

                return JsonResponse({'message': 'Data deleted successfully'}, status=status.HTTP_200_OK)
            else:
                return JsonResponse({'error': 'Sheet not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
